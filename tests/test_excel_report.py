"""Tests for the Excel report generator."""

from datetime import datetime
from decimal import Decimal
from pathlib import Path

import pytest
from openpyxl import load_workbook

from src.engine.models import (
    MatchConfidence,
    MatchResult,
    MatchStatus,
    ReconciliationSummary,
    Transaction,
    TransactionType,
)
from src.reports.excel_report import ExcelReportGenerator


def make_txn(
    id: str,
    date: str,
    amount: str,
    desc: str = "Test",
    source: str = "bank",
    ref: str = None,
) -> Transaction:
    """Helper to create test transactions."""
    amt = Decimal(amount)
    return Transaction(
        id=id,
        date=datetime.strptime(date, "%Y-%m-%d"),
        amount=amt,
        description=desc,
        type=TransactionType.CREDIT if amt >= 0 else TransactionType.DEBIT,
        reference=ref,
        source=source,
    )


@pytest.fixture
def sample_results():
    """Create sample reconciliation results for testing."""
    bank1 = make_txn("B1", "2025-01-10", "1000.00", "Payment A", "bank", "REF001")
    int1 = make_txn("I1", "2025-01-10", "1000.00", "Payment A", "internal", "REF001")

    bank2 = make_txn("B2", "2025-01-15", "500.00", "Payment B", "bank")
    int2 = make_txn("I2", "2025-01-16", "498.00", "Payment B fuzzy", "internal")

    bank3 = make_txn("B3", "2025-01-20", "-200.00", "Unmatched bank", "bank")
    int3 = make_txn("I3", "2025-01-25", "750.00", "Unmatched internal", "internal")

    bank4 = make_txn("B4", "2025-01-10", "1000.00", "Duplicate bank", "bank")

    results = [
        MatchResult(
            bank_transaction=bank1,
            internal_transaction=int1,
            status=MatchStatus.EXACT,
            confidence=MatchConfidence.HIGH,
            date_diff_days=0,
            amount_diff=Decimal("0"),
            match_reason="Exact match: same date, amount, and reference",
        ),
        MatchResult(
            bank_transaction=bank2,
            internal_transaction=int2,
            status=MatchStatus.FUZZY,
            confidence=MatchConfidence.MEDIUM,
            date_diff_days=1,
            amount_diff=Decimal("2.00"),
            match_reason="Fuzzy match: 1d date diff, 0.40% amount diff",
        ),
        MatchResult(
            bank_transaction=bank3,
            internal_transaction=None,
            status=MatchStatus.UNMATCHED_BANK,
            confidence=MatchConfidence.LOW,
            match_reason="No matching internal transaction found",
        ),
        MatchResult(
            bank_transaction=None,
            internal_transaction=int3,
            status=MatchStatus.UNMATCHED_INTERNAL,
            confidence=MatchConfidence.LOW,
            match_reason="No matching bank transaction found",
        ),
        MatchResult(
            bank_transaction=bank4,
            internal_transaction=None,
            status=MatchStatus.DUPLICATE,
            confidence=MatchConfidence.MEDIUM,
            match_reason="Potential duplicate: 2 bank transactions with same date and amount",
        ),
    ]
    return results


@pytest.fixture
def sample_summary():
    """Create sample reconciliation summary."""
    return ReconciliationSummary(
        total_bank_transactions=4,
        total_internal_transactions=3,
        total_matched=2,
        total_exact_matches=1,
        total_fuzzy_matches=1,
        total_unmatched_bank=1,
        total_unmatched_internal=1,
        total_duplicates=1,
        total_bank_amount=Decimal("2700.00"),
        total_internal_amount=Decimal("2248.00"),
        matched_amount=Decimal("1500.00"),
        unmatched_bank_amount=Decimal("200.00"),
        unmatched_internal_amount=Decimal("750.00"),
    )


class TestExcelReportGenerator:
    """Test Excel report generation functionality."""

    def test_generate_creates_file(self, tmp_path, sample_results, sample_summary):
        """Test that generate() creates an Excel file."""
        output = tmp_path / "test_report.xlsx"
        gen = ExcelReportGenerator()
        result_path = gen.generate(sample_results, sample_summary, output)

        assert result_path.exists()
        assert result_path.suffix == ".xlsx"

    def test_report_has_five_tabs(self, tmp_path, sample_results, sample_summary):
        """Test that the report has exactly 5 tabs."""
        output = tmp_path / "test_report.xlsx"
        gen = ExcelReportGenerator()
        gen.generate(sample_results, sample_summary, output)

        wb = load_workbook(output)
        assert len(wb.sheetnames) == 5
        assert "Summary" in wb.sheetnames
        assert "Matched" in wb.sheetnames
        assert "Bank Only" in wb.sheetnames
        assert "Internal Only" in wb.sheetnames
        assert "Duplicates" in wb.sheetnames

    def test_summary_tab_has_kpis(self, tmp_path, sample_results, sample_summary):
        """Test that Summary tab contains KPI data."""
        output = tmp_path / "test_report.xlsx"
        gen = ExcelReportGenerator()
        gen.generate(sample_results, sample_summary, output)

        wb = load_workbook(output)
        ws = wb["Summary"]

        # Title
        assert ws["A1"].value == "Bank Reconciliation Report"

        # KPIs should be present
        kpi_labels = []
        for row in range(5, 12):
            val = ws[f"A{row}"].value
            if val:
                kpi_labels.append(val)

        assert "Match Rate" in kpi_labels
        assert "Total Matched" in kpi_labels

    def test_matched_tab_has_data(self, tmp_path, sample_results, sample_summary):
        """Test that Matched tab contains matched transactions."""
        output = tmp_path / "test_report.xlsx"
        gen = ExcelReportGenerator()
        gen.generate(sample_results, sample_summary, output)

        wb = load_workbook(output)
        ws = wb["Matched"]

        # Header row
        assert ws["A1"].value == "Bank Date"
        assert ws["B1"].value == "Bank Amount"

        # Data rows (2 matched: 1 exact + 1 fuzzy)
        assert ws["A2"].value is not None  # First matched
        assert ws["A3"].value is not None  # Second matched

    def test_bank_only_tab(self, tmp_path, sample_results, sample_summary):
        """Test that Bank Only tab shows unmatched bank transactions."""
        output = tmp_path / "test_report.xlsx"
        gen = ExcelReportGenerator()
        gen.generate(sample_results, sample_summary, output)

        wb = load_workbook(output)
        ws = wb["Bank Only"]

        assert ws["A1"].value == "Date"
        # One unmatched bank transaction
        assert ws["A2"].value is not None

    def test_internal_only_tab(self, tmp_path, sample_results, sample_summary):
        """Test that Internal Only tab shows unmatched internal transactions."""
        output = tmp_path / "test_report.xlsx"
        gen = ExcelReportGenerator()
        gen.generate(sample_results, sample_summary, output)

        wb = load_workbook(output)
        ws = wb["Internal Only"]

        assert ws["A1"].value == "Date"
        # One unmatched internal transaction
        assert ws["A2"].value is not None

    def test_duplicates_tab(self, tmp_path, sample_results, sample_summary):
        """Test that Duplicates tab shows flagged duplicates."""
        output = tmp_path / "test_report.xlsx"
        gen = ExcelReportGenerator()
        gen.generate(sample_results, sample_summary, output)

        wb = load_workbook(output)
        ws = wb["Duplicates"]

        assert ws["A1"].value == "Source"
        # One duplicate
        assert ws["A2"].value is not None

    def test_empty_results(self, tmp_path):
        """Test report generation with empty results."""
        output = tmp_path / "empty_report.xlsx"
        summary = ReconciliationSummary()
        gen = ExcelReportGenerator()
        result_path = gen.generate([], summary, output)

        assert result_path.exists()
        wb = load_workbook(output)
        assert len(wb.sheetnames) == 5

    def test_output_directory_created(self, tmp_path, sample_results, sample_summary):
        """Test that output directory is created if it doesn't exist."""
        output = tmp_path / "subdir" / "nested" / "report.xlsx"
        gen = ExcelReportGenerator()
        result_path = gen.generate(sample_results, sample_summary, output)

        assert result_path.exists()

    def test_matched_tab_frozen_panes(self, tmp_path, sample_results, sample_summary):
        """Test that header rows are frozen."""
        output = tmp_path / "test_report.xlsx"
        gen = ExcelReportGenerator()
        gen.generate(sample_results, sample_summary, output)

        wb = load_workbook(output)
        ws = wb["Matched"]
        assert ws.freeze_panes == "A2"

    def test_number_formatting(self, tmp_path, sample_results, sample_summary):
        """Test that amount cells have proper number formatting."""
        output = tmp_path / "test_report.xlsx"
        gen = ExcelReportGenerator()
        gen.generate(sample_results, sample_summary, output)

        wb = load_workbook(output)
        ws = wb["Matched"]
        # B2 should have number format
        assert ws["B2"].number_format == '#,##0.00'
