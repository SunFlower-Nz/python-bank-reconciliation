"""Excel report generator for reconciliation results."""

from datetime import datetime
from decimal import Decimal
from pathlib import Path
from typing import List

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from src.engine.models import MatchConfidence, MatchResult, MatchStatus, ReconciliationSummary


class ExcelReportGenerator:
    """Generate professional Excel reports from reconciliation results."""

    # Style constants
    HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    HEADER_FONT = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    MATCHED_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    UNMATCHED_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    DUPLICATE_FILL = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    TITLE_FONT = Font(name="Calibri", size=16, bold=True, color="1F4E79")
    SUBTITLE_FONT = Font(name="Calibri", size=12, bold=True, color="1F4E79")
    KPI_FONT = Font(name="Calibri", size=14, bold=True)
    THIN_BORDER = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    def generate(
        self,
        results: List[MatchResult],
        summary: ReconciliationSummary,
        output_path: str | Path,
    ) -> Path:
        """
        Generate Excel report with 5 tabs.

        Args:
            results: List of match results from reconciliation.
            summary: Summary statistics.
            output_path: Path for the output Excel file.

        Returns:
            Path to the generated report.
        """
        output_path = Path(output_path)
        output_path.parent.mkdir(parents=True, exist_ok=True)

        wb = Workbook()

        # Tab 1: Summary
        self._create_summary_tab(wb, summary)

        # Tab 2: Matched Transactions
        matched = [r for r in results if r.is_matched]
        self._create_matched_tab(wb, matched)

        # Tab 3: Bank Only (Unmatched)
        bank_only = [r for r in results if r.status == MatchStatus.UNMATCHED_BANK]
        self._create_bank_only_tab(wb, bank_only)

        # Tab 4: Internal Only (Unmatched)
        internal_only = [r for r in results if r.status == MatchStatus.UNMATCHED_INTERNAL]
        self._create_internal_only_tab(wb, internal_only)

        # Tab 5: Duplicates
        duplicates = [r for r in results if r.status == MatchStatus.DUPLICATE]
        self._create_duplicates_tab(wb, duplicates)

        # Remove default sheet if extra
        if "Sheet" in wb.sheetnames and len(wb.sheetnames) > 1:
            del wb["Sheet"]

        wb.save(str(output_path))
        return output_path

    def _create_summary_tab(self, wb: Workbook, summary: ReconciliationSummary) -> None:
        """Create the Summary dashboard tab."""
        ws = wb.active
        ws.title = "Summary"
        ws.sheet_properties.tabColor = "1F4E79"

        # Title
        ws.merge_cells("A1:F1")
        ws["A1"] = "Bank Reconciliation Report"
        ws["A1"].font = self.TITLE_FONT
        ws["A1"].alignment = Alignment(horizontal="center")

        # Generated date
        ws.merge_cells("A2:F2")
        ws["A2"] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        ws["A2"].alignment = Alignment(horizontal="center")

        # KPIs
        kpis = [
            ("Match Rate", f"{summary.match_rate:.1f}%"),
            ("Total Matched", str(summary.total_matched)),
            ("Exact Matches", str(summary.total_exact_matches)),
            ("Fuzzy Matches", str(summary.total_fuzzy_matches)),
            ("Unmatched (Bank)", str(summary.total_unmatched_bank)),
            ("Unmatched (Internal)", str(summary.total_unmatched_internal)),
            ("Duplicates Found", str(summary.total_duplicates)),
        ]

        ws["A4"] = "Key Performance Indicators"
        ws["A4"].font = self.SUBTITLE_FONT

        for i, (label, value) in enumerate(kpis, start=5):
            ws[f"A{i}"] = label
            ws[f"A{i}"].font = Font(bold=True)
            ws[f"B{i}"] = value
            ws[f"B{i}"].font = self.KPI_FONT

            # Color coding
            if "Unmatched" in label and int(value) > 0:
                ws[f"B{i}"].fill = self.UNMATCHED_FILL
            elif "Match Rate" in label:
                rate = float(value.replace("%", ""))
                ws[f"B{i}"].fill = self.MATCHED_FILL if rate >= 95 else self.UNMATCHED_FILL

        # Amounts section
        row = len(kpis) + 7
        ws[f"A{row}"] = "Amount Summary"
        ws[f"A{row}"].font = self.SUBTITLE_FONT
        row += 1

        amounts = [
            ("Total Bank Amount", summary.total_bank_amount),
            ("Total Internal Amount", summary.total_internal_amount),
            ("Amount Difference", summary.amount_difference),
            ("Matched Amount", summary.matched_amount),
            ("Unmatched Bank Amount", summary.unmatched_bank_amount),
            ("Unmatched Internal Amount", summary.unmatched_internal_amount),
        ]

        for label, amount in amounts:
            ws[f"A{row}"] = label
            ws[f"A{row}"].font = Font(bold=True)
            ws[f"B{row}"] = float(amount)
            ws[f"B{row}"].number_format = '#,##0.00'
            row += 1

        # Auto-width
        ws.column_dimensions["A"].width = 25
        ws.column_dimensions["B"].width = 20

    def _create_matched_tab(self, wb: Workbook, matched: List[MatchResult]) -> None:
        """Create the Matched Transactions tab."""
        ws = wb.create_sheet("Matched")
        ws.sheet_properties.tabColor = "00B050"

        headers = [
            "Bank Date", "Bank Amount", "Bank Description", "Bank Ref",
            "Internal Date", "Internal Amount", "Internal Description", "Internal Ref",
            "Match Type", "Confidence", "Date Diff (days)", "Amount Diff", "Reason",
        ]

        self._write_headers(ws, headers)

        for i, result in enumerate(matched, start=2):
            bank = result.bank_transaction
            internal = result.internal_transaction

            ws[f"A{i}"] = bank.date.strftime("%Y-%m-%d") if bank else ""
            ws[f"B{i}"] = float(bank.amount) if bank else 0
            ws[f"B{i}"].number_format = '#,##0.00'
            ws[f"C{i}"] = (bank.description[:50] if bank else "")
            ws[f"D{i}"] = (bank.reference or "") if bank else ""
            ws[f"E{i}"] = internal.date.strftime("%Y-%m-%d") if internal else ""
            ws[f"F{i}"] = float(internal.amount) if internal else 0
            ws[f"F{i}"].number_format = '#,##0.00'
            ws[f"G{i}"] = (internal.description[:50]) if internal else ""
            ws[f"H{i}"] = (internal.reference or "") if internal else ""
            ws[f"I{i}"] = result.status.value
            ws[f"J{i}"] = result.confidence.value
            ws[f"K{i}"] = result.date_diff_days
            ws[f"L{i}"] = float(result.amount_diff)
            ws[f"L{i}"].number_format = '#,##0.00'
            ws[f"M{i}"] = result.match_reason

            # Apply row fill
            fill = self.MATCHED_FILL
            for col in range(1, len(headers) + 1):
                ws.cell(row=i, column=col).fill = fill

        self._auto_width(ws, headers)

    def _create_bank_only_tab(self, wb: Workbook, bank_only: List[MatchResult]) -> None:
        """Create the Bank Only tab."""
        ws = wb.create_sheet("Bank Only")
        ws.sheet_properties.tabColor = "FF0000"

        headers = ["Date", "Amount", "Description", "Reference", "Type", "Reason"]
        self._write_headers(ws, headers)

        for i, result in enumerate(bank_only, start=2):
            txn = result.bank_transaction
            if not txn:
                continue
            ws[f"A{i}"] = txn.date.strftime("%Y-%m-%d")
            ws[f"B{i}"] = float(txn.amount)
            ws[f"B{i}"].number_format = '#,##0.00'
            ws[f"C{i}"] = txn.description[:80]
            ws[f"D{i}"] = txn.reference or ""
            ws[f"E{i}"] = txn.type.value
            ws[f"F{i}"] = result.match_reason

            for col in range(1, len(headers) + 1):
                ws.cell(row=i, column=col).fill = self.UNMATCHED_FILL

        self._auto_width(ws, headers)

    def _create_internal_only_tab(self, wb: Workbook, internal_only: List[MatchResult]) -> None:
        """Create the Internal Only tab."""
        ws = wb.create_sheet("Internal Only")
        ws.sheet_properties.tabColor = "FF0000"

        headers = ["Date", "Amount", "Description", "Reference", "Type", "Reason"]
        self._write_headers(ws, headers)

        row = 2
        for result in internal_only:
            txn = result.internal_transaction
            if not txn:
                continue
            ws[f"A{row}"] = txn.date.strftime("%Y-%m-%d")
            ws[f"B{row}"] = float(txn.amount)
            ws[f"B{row}"].number_format = '#,##0.00'
            ws[f"C{row}"] = txn.description[:80]
            ws[f"D{row}"] = txn.reference or ""
            ws[f"E{row}"] = txn.type.value
            ws[f"F{row}"] = result.match_reason

            for col in range(1, len(headers) + 1):
                ws.cell(row=row, column=col).fill = self.UNMATCHED_FILL
            row += 1

        self._auto_width(ws, headers)

    def _create_duplicates_tab(self, wb: Workbook, duplicates: List[MatchResult]) -> None:
        """Create the Duplicates tab."""
        ws = wb.create_sheet("Duplicates")
        ws.sheet_properties.tabColor = "FFC000"

        headers = ["Source", "Date", "Amount", "Description", "Reference", "Reason"]
        self._write_headers(ws, headers)

        for i, result in enumerate(duplicates, start=2):
            txn = result.internal_transaction or result.bank_transaction
            source = "Internal" if result.internal_transaction else "Bank"

            ws[f"A{i}"] = source
            ws[f"B{i}"] = txn.date.strftime("%Y-%m-%d")
            ws[f"C{i}"] = float(txn.amount)
            ws[f"C{i}"].number_format = '#,##0.00'
            ws[f"D{i}"] = txn.description[:80]
            ws[f"E{i}"] = txn.reference or ""
            ws[f"F{i}"] = result.match_reason

            for col in range(1, len(headers) + 1):
                ws.cell(row=i, column=col).fill = self.DUPLICATE_FILL

        self._auto_width(ws, headers)

    def _write_headers(self, ws, headers: List[str]) -> None:
        """Write styled header row."""
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = self.HEADER_FONT
            cell.fill = self.HEADER_FILL
            cell.alignment = Alignment(horizontal="center")
            cell.border = self.THIN_BORDER

        # Freeze top row
        ws.freeze_panes = "A2"

        # Auto-filter
        ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"

    def _auto_width(self, ws, headers: List[str]) -> None:
        """Auto-adjust column widths."""
        for col_idx, header in enumerate(headers, start=1):
            col_letter = get_column_letter(col_idx)
            max_len = len(header) + 4
            ws.column_dimensions[col_letter].width = min(max_len, 35)
